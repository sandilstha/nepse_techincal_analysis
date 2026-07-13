"""
portfolio_views.py — private, per-user holdings portfolio + risk dashboard.

A logged-in user uploads a Meroshare "My Shares" export (CSV, Excel or PDF); we
parse it into their private ``Portfolio`` and render valuation / concentration /
sector-exposure / beta analytics (see ``services.portfolio_analytics``).
Everything here is gated behind login so one user can never see another's
positions. ``portfolio_sop_view`` serves the metric SOP that every (?) help
icon on the dashboard links into.
"""
from __future__ import annotations

import csv
import hashlib
import io
import logging
import math
import re
import zipfile

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.db.models import Count
from django.http import JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.views.decorators.http import require_POST

from core_analysis.forms import AdminApprovalRegistrationForm
from core_analysis.models import AccountApproval

logger = logging.getLogger(__name__)

DEFAULT_PORTFOLIO_NAME = "My Portfolio"
MAX_UPLOAD_BYTES = 2 * 1024 * 1024      # a holdings CSV is tiny; cap to be safe
MAX_DOC_UPLOAD_BYTES = 8 * 1024 * 1024  # holdings/WACC reports can be multi-page PDFs
MAX_EXCEL_UNCOMPRESSED_BYTES = 32 * 1024 * 1024
MAX_TABLE_ROWS = 5_000
MAX_TABLE_COLUMNS = 100
MAX_IMPORT_ROWS = 1_000
MAX_PDF_PAGES = 50
MAX_QUANTITY = 10**16
MAX_SNAPSHOT_PRICE = 10**12
MAX_WACC_RATE = 10**10
MAX_TOTAL_COST = 10**16
_ALLOWED_UPLOAD_EXTENSIONS = (".csv", ".pdf", ".xlsx", ".xlsm")
_DOC_EXTENSIONS = (".pdf", ".xlsx", ".xlsm")
_SYMBOL_RE = re.compile(r"^[A-Z0-9]+$")
_OWNER_NAME_RE = re.compile(
    r"(?:holder|bo|beneficiary(?:\s+owner)?|account\s+holder|client)\s+name"
    r"\s*[:\-]\s*(.+?)(?=\s+(?:boid|bo\s+id|client\s+id|dp\s+id|report\s+date|date)"
    r"\s*[:\-]|$)",
    re.IGNORECASE,
)


def _asset_version():
    from core_analysis.insights_views import _asset_version as v

    return v()


# ─────────────────────────────────────────────────────────────────────────────
# Holdings parsing (Meroshare "My Shares" export — CSV, Excel or PDF)
# ─────────────────────────────────────────────────────────────────────────────
def _num(value):
    """Parse a number from a CSV cell ('1,065.00', '"100"', '') -> float|None."""
    if value is None:
        return None
    s = str(value).replace(",", "").replace('"', "").strip()
    if not s:
        return None
    try:
        parsed = float(s)
    except (ValueError, OverflowError):
        return None
    return parsed if math.isfinite(parsed) else None


def _upload_validation_error(upload, report_name):
    """Return a user-facing error for unsupported or over-sized report files."""
    name = (getattr(upload, "name", "") or "").lower()
    extension = next(
        (suffix for suffix in _ALLOWED_UPLOAD_EXTENSIONS if name.endswith(suffix)),
        None,
    )
    if extension is None:
        return f"Unsupported {report_name} file. Upload a CSV, XLSX, XLSM or PDF file."
    limit = MAX_DOC_UPLOAD_BYTES if extension in _DOC_EXTENSIONS else MAX_UPLOAD_BYTES
    if upload.size and upload.size > limit:
        return f"That file is too large to be a {report_name} report."
    return None


def _check_table_shape(row_number, raw):
    """Reject documents whose expanded table shape is unreasonable for a portfolio."""
    if row_number > MAX_TABLE_ROWS:
        raise ValueError("Report contains too many table rows.")
    if len(raw or []) > MAX_TABLE_COLUMNS:
        raise ValueError("Report contains too many table columns.")


def _map_columns(header):
    """Locate the columns we need by header name (order-independent).

    The Meroshare export pairs each price with a rupee "Value as of …" column
    whose label *contains* the price label (e.g. "Value as of Last Closing
    Price"), so those value columns are skipped first and the first genuine match
    wins — otherwise the value column would shadow the price column.
    """
    col = {}
    for i, raw in enumerate(header):
        h = (raw or "").strip().lower()
        if not h or h.startswith("value") or "value as of" in h:
            continue  # rupee-value columns, never a price/quantity we want
        if h.startswith("scrip") and "symbol" not in col:
            col["symbol"] = i
        elif ("current balance" in h or h == "balance" or "quantity" in h) and "qty" not in col:
            col["qty"] = i
        elif "last closing price" in h and "close" not in col:
            col["close"] = i
        elif ("last transaction price" in h or h.endswith("(ltp)") or h == "ltp") and "ltp" not in col:
            col["ltp"] = i
    return col


def _normalize_holdings_table(table):
    """Map a raw table (list of cell-lists) to holdings rows + a skipped count.

    Works for CSV rows, Excel rows and pdfplumber-extracted PDF tables alike.
    Cells are whitespace-collapsed first (PDF headers carry embedded newlines).
    Tolerates the leading header row, the trailing ``Total :`` summary row,
    blank lines, quoted cells and missing decimals. Returns ``(rows, skipped)``
    where each row is ``{symbol, quantity, last_close, ltp}`` and ``skipped``
    counts non-position lines that were ignored (header/total/blank/invalid).
    """
    rows, skipped, col, seen = [], 0, None, set()
    for row_number, raw in enumerate(table, start=1):
        _check_table_shape(row_number, raw)
        norm = [" ".join(str(c or "").split()) for c in raw]
        if not norm or all(not c for c in norm):
            continue
        if col is None:
            if any("scrip" in c.lower() for c in norm):
                col = _map_columns(norm)
                continue
            skipped += 1
            continue
        if "symbol" not in col or col["symbol"] >= len(norm):
            skipped += 1
            continue
        sym = norm[col["symbol"]].strip().upper()
        if (
            not sym
            or len(sym) > 20
            or sym.startswith("TOTAL")
            or not _SYMBOL_RE.match(sym)
            or sym in seen
        ):
            skipped += 1
            continue
        cell = lambda k: norm[col[k]] if k in col and col[k] < len(norm) else None
        qty = _num(cell("qty"))
        if qty is None or not 0 < qty < MAX_QUANTITY:
            skipped += 1
            continue
        close = _num(cell("close"))
        ltp = _num(cell("ltp"))
        close = close if close is not None and 0 <= close < MAX_SNAPSHOT_PRICE else None
        ltp = ltp if ltp is not None and 0 <= ltp < MAX_SNAPSHOT_PRICE else None
        seen.add(sym)
        rows.append({
            "symbol": sym,
            "quantity": qty,
            "last_close": close,
            "ltp": ltp,
        })
        if len(rows) > MAX_IMPORT_ROWS:
            raise ValueError("Report contains too many holdings.")
    return rows, skipped


def parse_holdings_csv(text):
    """Back-compat CSV entry point — see ``_normalize_holdings_table``."""
    return _normalize_holdings_table(csv.reader(io.StringIO(text)))


def _extract_holdings_owner(table, document_text=""):
    """Return the account-holder name carried in a holdings report, if present."""
    sources = list((document_text or "").splitlines())
    sources.extend(
        " ".join(" ".join(str(cell or "").split()) for cell in row)
        for row in table
    )
    for source in sources:
        match = _OWNER_NAME_RE.search(" ".join(source.split()))
        if not match:
            continue
        candidate = match.group(1).strip(" :,|-")
        candidate = re.split(r"\s+(?=\d{8,}\b)", candidate, maxsplit=1)[0].strip()
        if not candidate or len(candidate) > 120 or not re.search(r"[A-Za-z]{2}", candidate):
            continue
        if candidate.isupper():
            candidate = candidate.title()
        return _normalize_portfolio_name(candidate)
    return ""


def parse_holdings_report_details(upload):
    """Parse holdings plus metadata -> ``(rows, skipped, owner_name)``."""
    name = (upload.name or "").lower()
    data = upload.read()
    if name.endswith(".pdf"):
        table, document_text = _table_and_text_from_pdf(data)
    elif name.endswith((".xlsx", ".xlsm")):
        table = _table_from_excel(data)
        document_text = ""
    else:
        document_text = data.decode("utf-8-sig", errors="replace")
        table = list(csv.reader(io.StringIO(document_text)))
    rows, skipped = _normalize_holdings_table(table)
    return rows, skipped, _extract_holdings_owner(table, document_text)


def parse_holdings_report(upload):
    """Back-compatible holdings parser returning ``(rows, skipped)``."""
    rows, skipped, _owner = parse_holdings_report_details(upload)
    return rows, skipped


# ─────────────────────────────────────────────────────────────────────────────
# WACC / cost-basis parsing (broker "My WACC" report — CSV, Excel or PDF)
# ─────────────────────────────────────────────────────────────────────────────
def _map_wacc_columns(header):
    """Locate the WACC report's columns by header name (order-independent).

    Header labels may carry embedded newlines in the PDF ("WACC\\nCalculated\\n
    Quantity"); callers pass whitespace-collapsed cells so the substring match
    below is stable across CSV / Excel / PDF exports.
    """
    col = {}
    for i, raw in enumerate(header):
        h = (raw or "").strip().lower()
        if h.startswith("scrip") and "symbol" not in col:
            col["symbol"] = i
        elif "wacc rate" in h and "rate" not in col:
            col["rate"] = i
        elif "calculated quantity" in h and "qty" not in col:
            col["qty"] = i
        elif "total cost" in h and "cost" not in col:
            col["cost"] = i
        elif "modification" in h and "modified" not in col:
            col["modified"] = i
    return col


def _normalize_wacc_table(table):
    """Map a raw table (list of cell-lists) to WACC rows + a skipped count.

    Returns ``(rows, skipped)`` where each row is
    ``{symbol, wacc_rate, quantity, total_cost, modified}``. Rows without a
    usable WACC rate, the header, and any total line are skipped.
    """
    rows, skipped, col, seen = [], 0, None, set()
    for row_number, raw in enumerate(table, start=1):
        _check_table_shape(row_number, raw)
        norm = [" ".join(str(c or "").split()) for c in raw]
        if not norm or all(not c for c in norm):
            continue
        if col is None:
            if any("scrip" in c.lower() for c in norm):
                col = _map_wacc_columns(norm)
                continue
            skipped += 1
            continue
        if "symbol" not in col or col["symbol"] >= len(norm):
            skipped += 1
            continue
        sym = norm[col["symbol"]].strip().upper()
        if (
            not sym
            or len(sym) > 20
            or sym.startswith("TOTAL")
            or not _SYMBOL_RE.match(sym)
            or sym in seen
        ):
            skipped += 1
            continue
        cell = lambda k: norm[col[k]] if k in col and col[k] < len(norm) else None
        rate, cost = _num(cell("rate")), _num(cell("cost"))
        qty = _num(cell("qty"))
        if rate is None:
            skipped += 1
            continue
        if not 0 <= rate < MAX_WACC_RATE:
            skipped += 1
            continue
        if qty is not None and not 0 <= qty < MAX_QUANTITY:
            skipped += 1
            continue
        if cost is not None and not 0 <= cost < MAX_TOTAL_COST:
            skipped += 1
            continue
        seen.add(sym)
        rows.append({
            "symbol": sym,
            "wacc_rate": rate,
            "quantity": qty,
            "total_cost": cost,
            "modified": (cell("modified") or "")[:32],
        })
        if len(rows) > MAX_IMPORT_ROWS:
            raise ValueError("Report contains too many WACC rows.")
    return rows, skipped


def _table_and_text_from_pdf(data):
    """All PDF table rows and page text, in document order."""
    import pdfplumber

    out, text = [], []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        if len(pdf.pages) > MAX_PDF_PAGES:
            raise ValueError("PDF contains too many pages.")
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text.append(page_text)
            for tbl in page.extract_tables() or ():
                if not tbl:
                    continue
                if len(out) + len(tbl) > MAX_TABLE_ROWS:
                    raise ValueError("PDF contains too many table rows.")
                for offset, row in enumerate(tbl, start=1):
                    _check_table_shape(len(out) + offset, row)
                out.extend(tbl)
    return out, "\n".join(text)


def _table_from_pdf(data):
    """All table rows from every page of a PDF, in document order."""
    table, _text = _table_and_text_from_pdf(data)
    return table


def _table_from_excel(data):
    from openpyxl import load_workbook

    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        members = archive.infolist()
        if len(members) > 1_000:
            raise ValueError("Spreadsheet archive contains too many files.")
        if sum(member.file_size for member in members) > MAX_EXCEL_UNCOMPRESSED_BYTES:
            raise ValueError("Spreadsheet expands beyond the safe processing limit.")
    wb = load_workbook(
        io.BytesIO(data), read_only=True, data_only=True, keep_links=False
    )
    try:
        ws = wb.active
        if ws.max_row > MAX_TABLE_ROWS or ws.max_column > MAX_TABLE_COLUMNS:
            raise ValueError("Spreadsheet dimensions exceed the safe processing limit.")
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def parse_wacc_report(upload):
    """Parse an uploaded broker 'My WACC' report (CSV / Excel / PDF) -> rows.

    Dispatches on the filename extension; every branch yields a raw table that
    ``_normalize_wacc_table`` maps to ``{symbol, wacc_rate, quantity,
    total_cost, modified}`` rows.
    """
    name = (upload.name or "").lower()
    data = upload.read()
    if name.endswith(".pdf"):
        table = _table_from_pdf(data)
    elif name.endswith((".xlsx", ".xlsm")):
        table = _table_from_excel(data)
    else:
        text = data.decode("utf-8-sig", errors="replace")
        table = list(csv.reader(io.StringIO(text)))
    return _normalize_wacc_table(table)


def parse_broker_ledger_report(upload):
    """Parse a broker Account Statement (PDF/CSV/Excel) and hash its content."""
    from core_analysis.services.portfolio_ledger import normalize_broker_ledger

    name = (upload.name or "").lower()
    data = upload.read()
    file_sha256 = hashlib.sha256(data).hexdigest()
    document_text = ""
    if name.endswith(".pdf"):
        table, document_text = _table_and_text_from_pdf(data)
    elif name.endswith((".xlsx", ".xlsm")):
        table = _table_from_excel(data)
    else:
        document_text = data.decode("utf-8-sig", errors="replace")
        table = list(csv.reader(io.StringIO(document_text)))
    return normalize_broker_ledger(table, document_text), file_sha256


# ─────────────────────────────────────────────────────────────────────────────
# Auth
# ─────────────────────────────────────────────────────────────────────────────
def register_view(request):
    """Self-service signup that waits for admin approval."""
    if request.user.is_authenticated:
        return redirect("portfolio")
    if request.method == "POST":
        form = AdminApprovalRegistrationForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                user = form.save()
                AccountApproval.objects.create(
                    user=user, contact_email=form.cleaned_data["email"]
                )
            messages.success(
                request,
                "Account request sent to admin. You can sign in after approval.",
            )
            return redirect("approval_pending")
    else:
        form = AdminApprovalRegistrationForm()
    return render(
        request,
        "registration/register.html",
        {"form": form, "asset_version": _asset_version()},
    )


def approval_pending_view(request):
    return render(
        request,
        "registration/approval_pending.html",
        {"asset_version": _asset_version()},
    )


# ─────────────────────────────────────────────────────────────────────────────
# Portfolio page + data
# ─────────────────────────────────────────────────────────────────────────────
def _coerce_id(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _normalize_portfolio_name(value):
    return " ".join(str(value or "").split())[:120]


def _selected_portfolio_id(request):
    return (
        _coerce_id(request.POST.get("portfolio_id"))
        or _coerce_id(request.POST.get("portfolio"))
        or _coerce_id(request.GET.get("portfolio"))
        or _coerce_id(request.session.get("active_portfolio_id"))
    )


def _portfolio_url(portfolio=None):
    base = reverse("portfolio")
    return f"{base}?portfolio={portfolio.id}" if portfolio else base


def _remember_portfolio(request, portfolio):
    request.session["active_portfolio_id"] = portfolio.id


def _ensure_portfolio_defaults(user):
    """Guarantee that each account has one active default portfolio."""
    from core_analysis.models import Portfolio

    with transaction.atomic():
        portfolios = list(
            Portfolio.objects.select_for_update().filter(user=user).order_by("id")
        )
        if not portfolios:
            return Portfolio.objects.create(
                user=user,
                name=DEFAULT_PORTFOLIO_NAME,
                portfolio_type=Portfolio.PERSONAL,
                is_default=True,
            )

        active = [p for p in portfolios if not p.is_archived]
        if not active:
            selected = sorted(portfolios, key=lambda p: p.updated_at, reverse=True)[0]
            Portfolio.objects.filter(pk=selected.pk).update(is_archived=False, is_default=True)
            Portfolio.objects.filter(user=user).exclude(pk=selected.pk).update(is_default=False)
            selected.refresh_from_db()
            return selected

        defaults = [p for p in active if p.is_default]
        keep = sorted(defaults or active, key=lambda p: p.updated_at, reverse=True)[0]
        Portfolio.objects.filter(user=user, is_default=True).exclude(pk=keep.pk).update(is_default=False)
        if not keep.is_default:
            Portfolio.objects.filter(pk=keep.pk).update(is_default=True)
            keep.is_default = True
        keep.refresh_from_db()
        return keep


def _owned_portfolio(user, portfolio_id, *, include_archived=False):
    from core_analysis.models import Portfolio

    pid = _coerce_id(portfolio_id)
    if not pid:
        return None
    qs = Portfolio.objects.filter(user=user)
    if not include_archived:
        qs = qs.filter(is_archived=False)
    return qs.filter(pk=pid).first()


def _get_selected_portfolio(request, *, include_archived=False):
    from core_analysis.models import Portfolio

    default = _ensure_portfolio_defaults(request.user)
    portfolio = _owned_portfolio(
        request.user, _selected_portfolio_id(request), include_archived=include_archived
    )
    if portfolio and (include_archived or not portfolio.is_archived):
        if not portfolio.is_archived:
            _remember_portfolio(request, portfolio)
        return portfolio

    qs = Portfolio.objects.filter(user=request.user)
    if not include_archived:
        qs = qs.filter(is_archived=False)
    portfolio = qs.filter(is_default=True).first() or qs.order_by("-updated_at").first() or default
    if not portfolio.is_archived:
        _remember_portfolio(request, portfolio)
    return portfolio


def _get_mutation_portfolio(request):
    """Resolve a posted target strictly; never fall back after an invalid explicit ID."""
    raw_id = request.POST.get("portfolio_id")
    if raw_id in (None, ""):
        raw_id = request.POST.get("portfolio")
    if raw_id not in (None, ""):
        return _owned_portfolio(request.user, raw_id)
    return _get_selected_portfolio(request)


def _locked_active_portfolio(user, portfolio_id):
    """Lock an owned active portfolio before replacing its imported rows."""
    from core_analysis.models import Portfolio

    return (
        Portfolio.objects.select_for_update()
        .filter(user=user, pk=portfolio_id, is_archived=False)
        .first()
    )


def _unique_portfolio_name(user, base_name):
    from core_analysis.models import Portfolio

    base = _normalize_portfolio_name(base_name) or DEFAULT_PORTFOLIO_NAME
    existing = {
        name.casefold()
        for name in Portfolio.objects.filter(user=user).values_list("name", flat=True)
    }
    if base.casefold() not in existing:
        return base
    for n in range(2, 1000):
        suffix = f" {n}"
        candidate = f"{base[:120 - len(suffix)]}{suffix}"
        if candidate.casefold() not in existing:
            return candidate
    return f"{base[:112]} {Portfolio.objects.filter(user=user).count() + 1}"


def _portfolio_rows(user):
    from core_analysis.models import Portfolio

    rows = (
        Portfolio.objects.filter(user=user)
        .annotate(
            holdings_count=Count("holdings", distinct=True),
            costs_count=Count("costs", distinct=True),
            ledger_count=Count("ledger_transactions", distinct=True),
        )
        .order_by("is_archived", "-is_default", "name")
    )
    active, archived = [], []
    for row in rows:
        (archived if row.is_archived else active).append(row)
    return active, archived


def _is_approver(user):
    """Who may review account requests from the portfolio notification bell.

    Staff (the admin site's own gate) — this is the "main admin" account that
    approves signups. Kept as one predicate so the view, the API and the action
    endpoint can never drift apart.
    """
    return bool(user.is_authenticated and user.is_staff)


@login_required(login_url="login")
def portfolio_view(request):
    """Render the Risk & Portfolio dashboard shell for the logged-in user."""
    from core_analysis.services.portfolio_ledger import build_ledger_dashboard

    portfolio = _get_selected_portfolio(request)
    ledger = build_ledger_dashboard(
        portfolio,
        fiscal_year=request.GET.get("fy"),
        row_limit=5,
        page=request.GET.get("ledger_page"),
    )
    active_portfolios, archived_portfolios = _portfolio_rows(request.user)
    is_approver = _is_approver(request.user)
    pending_count = 0
    if is_approver:
        from core_analysis.models import AccountApproval as _AA

        pending_count = _AA.objects.filter(status=_AA.PENDING).count()
    return render(
        request,
        "core_analysis/portfolio.html",
        {
            "asset_version": _asset_version(),
            "active_portfolio": portfolio,
            "active_portfolios": active_portfolios,
            "archived_portfolios": archived_portfolios,
            "has_holdings": portfolio.holdings.exists(),
            "ledger": ledger,
            "is_approver": is_approver,
            "pending_approvals": pending_count,
        },
    )


@login_required(login_url="login")
def portfolio_sop_view(request):
    """Methodology SOP for the Risk & Portfolio desk.

    One anchored section per metric/card; every (?) help icon on the dashboard
    deep-links to its section here. Static content — all formulas, assumptions
    and data sources are documented in the template itself.
    """
    return render(
        request,
        "core_analysis/portfolio_sop.html",
        {"asset_version": _asset_version()},
    )


@login_required(login_url="login")
def portfolio_data_api(request):
    """JSON valuation + risk roll-up for the user's portfolio."""
    from core_analysis.services import portfolio_analytics as pa

    requested_id = request.GET.get("portfolio")
    if requested_id not in (None, ""):
        portfolio = _owned_portfolio(request.user, requested_id)
        if portfolio is None:
            return JsonResponse({"ok": False, "error": "Portfolio not found."}, status=404)
        _remember_portfolio(request, portfolio)
    else:
        portfolio = _get_selected_portfolio(request)
    try:
        build_kwargs = {
            "participation_rate": request.GET.get("participation", pa.PARTICIPATION_RATE),
            "liquidation_target": request.GET.get("liquidation_pct", 100.0),
        }
        fiscal_year = (request.GET.get("fy") or "").strip()
        if fiscal_year:
            build_kwargs["fiscal_year"] = fiscal_year
        payload = pa.build_portfolio_payload(portfolio, **build_kwargs)
        payload.update({
            "portfolio_id": portfolio.id,
            "portfolio_name": portfolio.name,
            "portfolio_type": portfolio.portfolio_type,
            "portfolio_type_label": portfolio.get_portfolio_type_display(),
            "portfolio_is_default": portfolio.is_default,
        })
        return JsonResponse(payload)
    except Exception:  # pragma: no cover - defensive, never 500 the dashboard
        logger.exception(
            "portfolio payload failed for user %s portfolio %s",
            request.user.id,
            portfolio.id,
        )
        return JsonResponse({"ok": False, "error": "Unable to compute portfolio."}, status=500)


@login_required(login_url="login")
@require_POST
def portfolio_manage(request):
    """Create, rename, duplicate, archive, restore, delete and set defaults."""
    from core_analysis.models import Holding, HoldingCost, Portfolio

    action = (request.POST.get("action") or "").strip().lower()
    redirect_portfolio = None
    create_import = None
    create_cost_import = None

    # Parse an optional holdings file before opening the database transaction.
    # Selecting a file on the create form therefore creates, names and fills the
    # portfolio in one request; an invalid file never leaves an empty portfolio.
    if action == "create" and request.FILES.get("file"):
        upload = request.FILES["file"]
        upload_error = _upload_validation_error(upload, "holdings")
        if upload_error:
            messages.error(request, upload_error)
            return redirect(_portfolio_url(_get_selected_portfolio(request)))
        try:
            create_import = parse_holdings_report_details(upload)
        except Exception:
            logger.exception("holdings parse failed while creating portfolio for user %s", request.user.id)
            messages.error(
                request,
                "Could not read that file — upload the Meroshare 'My Shares' export.",
            )
            return redirect(_portfolio_url(_get_selected_portfolio(request)))
        if not create_import[0]:
            messages.error(request, "No holdings found in that file. Check the format and try again.")
            return redirect(_portfolio_url(_get_selected_portfolio(request)))

    if action == "create" and request.FILES.get("wacc_file"):
        wacc_upload = request.FILES["wacc_file"]
        upload_error = _upload_validation_error(wacc_upload, "WACC")
        if upload_error:
            messages.error(request, upload_error)
            return redirect(_portfolio_url(_get_selected_portfolio(request)))
        try:
            create_cost_import = parse_wacc_report(wacc_upload)
        except Exception:
            logger.exception("WACC parse failed while creating portfolio for user %s", request.user.id)
            messages.error(
                request,
                "Could not read that file — upload the broker 'My WACC' report.",
            )
            return redirect(_portfolio_url(_get_selected_portfolio(request)))
        if not create_cost_import[0]:
            messages.error(request, "No WACC rows found in that file. Check the format and try again.")
            return redirect(_portfolio_url(_get_selected_portfolio(request)))

    try:
        with transaction.atomic():
            default_portfolio = _ensure_portfolio_defaults(request.user)

            if action == "create":
                name = _normalize_portfolio_name(request.POST.get("name"))
                ptype = request.POST.get("portfolio_type") or Portfolio.CUSTOM
                valid_types = {key for key, _label in Portfolio.TYPE_CHOICES}
                if ptype not in valid_types:
                    ptype = Portfolio.CUSTOM
                if create_import:
                    rows, skipped, owner_name = create_import
                    type_label = dict(Portfolio.TYPE_CHOICES).get(ptype, DEFAULT_PORTFOLIO_NAME)
                    if not name:
                        name = _unique_portfolio_name(
                            request.user, owner_name or type_label
                        )
                if not name:
                    # No name field on the form — derive a unique one from the type
                    # (e.g. "Personal", then "Personal 2"). Rename afterwards to taste.
                    type_label = dict(Portfolio.TYPE_CHOICES).get(ptype, DEFAULT_PORTFOLIO_NAME)
                    name = _unique_portfolio_name(request.user, type_label)
                elif Portfolio.objects.filter(user=request.user, name__iexact=name).exists():
                    messages.error(request, "A portfolio with that name already exists.")
                    return redirect(_portfolio_url(_get_selected_portfolio(request)))
                reuse_placeholder = (
                    (create_import or create_cost_import)
                    and Portfolio.objects.filter(user=request.user).count() == 1
                    and default_portfolio.name == DEFAULT_PORTFOLIO_NAME
                    and not default_portfolio.holdings.exists()
                    and not default_portfolio.costs.exists()
                )
                if reuse_placeholder:
                    redirect_portfolio = default_portfolio
                    redirect_portfolio.name = name
                    redirect_portfolio.portfolio_type = ptype
                    redirect_portfolio.save(
                        update_fields=["name", "portfolio_type", "updated_at"]
                    )
                else:
                    make_default = not Portfolio.objects.filter(
                        user=request.user, is_archived=False
                    ).exists()
                    redirect_portfolio = Portfolio.objects.create(
                        user=request.user,
                        name=name,
                        portfolio_type=ptype,
                        is_default=make_default,
                    )
                if create_import:
                    Holding.objects.bulk_create([
                        Holding(
                            portfolio=redirect_portfolio,
                            symbol=row["symbol"],
                            quantity=row["quantity"],
                            last_close=row["last_close"],
                            ltp=row["ltp"],
                        )
                        for row in rows
                    ])
                    note = (
                        f"Created portfolio '{redirect_portfolio.name}' and imported "
                        f"{len(rows)} holdings."
                    )
                    if not owner_name and not _normalize_portfolio_name(request.POST.get("name")):
                        note += " Owner name was not present in the report; you can rename it."
                    if skipped:
                        note += f" ({skipped} non-position lines skipped.)"
                    if create_cost_import:
                        cost_rows, cost_skipped = create_cost_import
                        HoldingCost.objects.bulk_create([
                            HoldingCost(
                                portfolio=redirect_portfolio,
                                symbol=row["symbol"],
                                wacc_rate=row["wacc_rate"],
                                quantity=row["quantity"] or 0,
                                total_cost=row["total_cost"],
                                modified=row["modified"] or "",
                            )
                            for row in cost_rows
                        ])
                        note += f" Imported {len(cost_rows)} WACC rows."
                        if cost_skipped:
                            note += f" ({cost_skipped} non-data WACC lines skipped.)"
                    messages.success(request, note)
                elif create_cost_import:
                    cost_rows, cost_skipped = create_cost_import
                    HoldingCost.objects.bulk_create([
                        HoldingCost(
                            portfolio=redirect_portfolio,
                            symbol=row["symbol"],
                            wacc_rate=row["wacc_rate"],
                            quantity=row["quantity"] or 0,
                            total_cost=row["total_cost"],
                            modified=row["modified"] or "",
                        )
                        for row in cost_rows
                    ])
                    note = (
                        f"Created portfolio '{redirect_portfolio.name}' and imported "
                        f"{len(cost_rows)} WACC rows."
                    )
                    if cost_skipped:
                        note += f" ({cost_skipped} non-data lines skipped.)"
                    messages.success(request, note)
                else:
                    messages.success(request, f"Created portfolio '{redirect_portfolio.name}'.")

            elif action == "rename":
                portfolio = _owned_portfolio(
                    request.user, request.POST.get("portfolio_id"), include_archived=True
                )
                name = _normalize_portfolio_name(request.POST.get("name"))
                if portfolio is None or not name:
                    messages.error(request, "Choose a portfolio and enter a valid name.")
                    return redirect(_portfolio_url(_get_selected_portfolio(request)))
                if (
                    Portfolio.objects.filter(user=request.user, name__iexact=name)
                    .exclude(pk=portfolio.pk)
                    .exists()
                ):
                    messages.error(request, "A portfolio with that name already exists.")
                    return redirect(_portfolio_url(portfolio if not portfolio.is_archived else None))
                portfolio.name = name
                portfolio.save(update_fields=["name", "updated_at"])
                redirect_portfolio = portfolio if not portfolio.is_archived else _get_selected_portfolio(request)
                messages.success(request, f"Renamed portfolio to '{portfolio.name}'.")

            elif action == "duplicate":
                source = _owned_portfolio(
                    request.user, request.POST.get("portfolio_id"), include_archived=True
                )
                if source is None:
                    messages.error(request, "Portfolio not found.")
                    return redirect(_portfolio_url(_get_selected_portfolio(request)))
                name = _unique_portfolio_name(request.user, f"{source.name} Copy")
                redirect_portfolio = Portfolio.objects.create(
                    user=request.user,
                    name=name,
                    portfolio_type=source.portfolio_type,
                    is_default=False,
                    is_archived=False,
                )
                Holding.objects.bulk_create([
                    Holding(
                        portfolio=redirect_portfolio,
                        symbol=h.symbol,
                        quantity=h.quantity,
                        last_close=h.last_close,
                        ltp=h.ltp,
                    )
                    for h in source.holdings.all()
                ])
                HoldingCost.objects.bulk_create([
                    HoldingCost(
                        portfolio=redirect_portfolio,
                        symbol=c.symbol,
                        wacc_rate=c.wacc_rate,
                        quantity=c.quantity,
                        total_cost=c.total_cost,
                        modified=c.modified,
                    )
                    for c in source.costs.all()
                ])
                from core_analysis.services.portfolio_ledger import duplicate_ledger

                duplicate_ledger(source, redirect_portfolio)
                messages.success(request, f"Duplicated '{source.name}' as '{redirect_portfolio.name}'.")

            elif action == "set_default":
                portfolio = _owned_portfolio(request.user, request.POST.get("portfolio_id"))
                if portfolio is None:
                    messages.error(request, "Archived portfolios cannot be set as default.")
                    return redirect(_portfolio_url(_get_selected_portfolio(request)))
                Portfolio.objects.filter(user=request.user).update(is_default=False)
                portfolio.is_default = True
                portfolio.save(update_fields=["is_default", "updated_at"])
                redirect_portfolio = portfolio
                messages.success(request, f"'{portfolio.name}' is now your default portfolio.")

            elif action == "archive":
                portfolio = _owned_portfolio(request.user, request.POST.get("portfolio_id"))
                if portfolio is None:
                    messages.error(request, "Portfolio not found.")
                    return redirect(_portfolio_url(_get_selected_portfolio(request)))
                active_count = Portfolio.objects.filter(user=request.user, is_archived=False).count()
                if active_count <= 1:
                    messages.error(request, "Create another active portfolio before archiving this one.")
                    return redirect(_portfolio_url(portfolio))
                portfolio.is_archived = True
                portfolio.is_default = False
                portfolio.save(update_fields=["is_archived", "is_default", "updated_at"])
                redirect_portfolio = _ensure_portfolio_defaults(request.user)
                messages.success(request, f"Archived '{portfolio.name}'.")

            elif action == "restore":
                portfolio = _owned_portfolio(
                    request.user, request.POST.get("portfolio_id"), include_archived=True
                )
                if portfolio is None:
                    messages.error(request, "Portfolio not found.")
                    return redirect(_portfolio_url(_get_selected_portfolio(request)))
                portfolio.is_archived = False
                portfolio.save(update_fields=["is_archived", "updated_at"])
                redirect_portfolio = portfolio
                _ensure_portfolio_defaults(request.user)
                messages.success(request, f"Restored '{portfolio.name}'.")

            elif action == "delete":
                portfolio = _owned_portfolio(
                    request.user, request.POST.get("portfolio_id"), include_archived=True
                )
                if portfolio is None:
                    messages.error(request, "Portfolio not found.")
                    return redirect(_portfolio_url(_get_selected_portfolio(request)))
                if not portfolio.is_archived:
                    active_count = (
                        Portfolio.objects.select_for_update()
                        .filter(user=request.user, is_archived=False)
                        .count()
                    )
                    if active_count <= 1:
                        messages.error(
                            request,
                            "The last active portfolio cannot be deleted. Create another portfolio first.",
                        )
                        return redirect(_portfolio_url(portfolio))
                name = portfolio.name
                portfolio.delete()
                redirect_portfolio = _ensure_portfolio_defaults(request.user)
                messages.success(request, f"Deleted portfolio '{name}'.")

            else:
                messages.error(request, "Unknown portfolio action.")
                return redirect(_portfolio_url(_get_selected_portfolio(request)))

    except IntegrityError:
        logger.exception("portfolio management action failed for user %s", request.user.id)
        messages.error(request, "That portfolio change conflicts with an existing portfolio.")
        return redirect(_portfolio_url(_get_selected_portfolio(request)))

    if redirect_portfolio and not redirect_portfolio.is_archived:
        _remember_portfolio(request, redirect_portfolio)
    return redirect(_portfolio_url(redirect_portfolio if redirect_portfolio and not redirect_portfolio.is_archived else None))


# ─────────────────────────────────────────────────────────────────────────────
# Account-approval notifications (admin only, surfaced on the portfolio page)
# ─────────────────────────────────────────────────────────────────────────────
def _approval_json(a):
    """One pending request → the JSON the notification bell renders."""
    return {
        "id": a.id,
        "username": a.user.username,
        "email": a.contact_email or (a.user.email or ""),
        "requested_at": a.created_at.isoformat() if a.created_at else None,
    }


@login_required(login_url="login")
def portfolio_approvals_api(request):
    """Pending account requests for the notification bell (staff only).

    Returns ``{ok, is_approver, pending_count, requests:[…]}``. Non-staff get a
    quiet ``is_approver: False`` (not a 403) so the bell simply never shows for
    ordinary users without the frontend having to special-case anything.
    """
    from core_analysis.models import AccountApproval

    if not _is_approver(request.user):
        return JsonResponse({"ok": True, "is_approver": False, "pending_count": 0, "requests": []})

    pending = list(
        AccountApproval.objects.filter(status=AccountApproval.PENDING)
        .select_related("user")
        .order_by("-created_at")
    )
    return JsonResponse({
        "ok": True,
        "is_approver": True,
        "pending_count": len(pending),
        "requests": [_approval_json(a) for a in pending],
    })


@login_required(login_url="login")
@require_POST
def portfolio_approval_action(request):
    """Approve or reject one pending account request from the bell (staff only).

    Body: ``id`` + ``action`` in {approve, reject}. Activates/deactivates the
    user via the model's own ``approve``/``reject`` (same path as the admin
    action), stamps this admin as reviewer, and returns the fresh pending list
    so the panel and badge update in one round-trip.
    """
    from core_analysis.models import AccountApproval

    if not _is_approver(request.user):
        return JsonResponse({"ok": False, "error": "Not authorised."}, status=403)

    action = (request.POST.get("action") or "").strip().lower()
    if action not in ("approve", "reject"):
        return JsonResponse({"ok": False, "error": "Unknown action."}, status=400)
    try:
        approval_id = int(request.POST.get("id") or 0)
    except (TypeError, ValueError):
        approval_id = 0

    approval = (
        AccountApproval.objects.select_related("user").filter(id=approval_id).first()
        if approval_id else None
    )
    if approval is None:
        return JsonResponse({"ok": False, "error": "Request not found."}, status=404)

    already = approval.status != AccountApproval.PENDING
    if not already:
        try:
            if action == "approve":
                approval.approve(reviewer=request.user)
            else:
                approval.reject(reviewer=request.user)
        except Exception:  # pragma: no cover - never 500 the bell
            logger.exception("approval action failed (id=%s) by user %s", approval_id, request.user.id)
            return JsonResponse({"ok": False, "error": "Could not update that request."}, status=500)

    remaining = list(
        AccountApproval.objects.filter(status=AccountApproval.PENDING)
        .select_related("user")
        .order_by("-created_at")
    )
    verb = "approved" if action == "approve" else "rejected"
    return JsonResponse({
        "ok": True,
        "email": approval.contact_email or (approval.user.email or ""),
        "username": approval.user.username,
        "result": approval.status,
        "message": (f"Already {approval.get_status_display().lower()}."
                    if already else f"{approval.user.username} {verb}."),
        "pending_count": len(remaining),
        "requests": [_approval_json(a) for a in remaining],
    })


@login_required(login_url="login")
@require_POST
def portfolio_import(request):
    """Replace the user's holdings with an uploaded 'My Shares' file (CSV / Excel / PDF)."""
    from core_analysis.models import Holding, Portfolio

    portfolio = _get_mutation_portfolio(request)
    if portfolio is None:
        messages.error(request, "Portfolio not found.")
        return redirect(_portfolio_url(_get_selected_portfolio(request)))

    upload = request.FILES.get("file")
    if not upload:
        messages.error(request, "Please choose a holdings file (CSV, Excel or PDF) to upload.")
        return redirect(_portfolio_url(portfolio))
    upload_error = _upload_validation_error(upload, "holdings")
    if upload_error:
        messages.error(request, upload_error)
        return redirect(_portfolio_url(portfolio))

    try:
        rows, skipped, owner_name = parse_holdings_report_details(upload)
    except Exception:
        logger.exception("holdings parse failed for user %s", request.user.id)
        messages.error(
            request,
            "Could not read that file — upload the Meroshare 'My Shares' export (CSV, Excel or PDF).",
        )
        return redirect(_portfolio_url(portfolio))

    if not rows:
        messages.error(request, "No holdings found in that file. Check the format and try again.")
        return redirect(_portfolio_url(portfolio))

    with transaction.atomic():
        portfolio = _locked_active_portfolio(request.user, portfolio.id)
        if portfolio is None:
            messages.error(request, "Portfolio not found or no longer active.")
            return redirect(_portfolio_url(_get_selected_portfolio(request)))
        auto_named = False
        if (
            owner_name
            and portfolio.name == DEFAULT_PORTFOLIO_NAME
            and not portfolio.holdings.exists()
            and not portfolio.costs.exists()
            and not Portfolio.objects.filter(
                user=request.user, name__iexact=owner_name
            ).exclude(pk=portfolio.pk).exists()
        ):
            portfolio.name = owner_name
            auto_named = True
        portfolio.holdings.all().delete()
        Holding.objects.bulk_create([
            Holding(
                portfolio=portfolio,
                symbol=r["symbol"],
                quantity=r["quantity"],
                last_close=r["last_close"],
                ltp=r["ltp"],
            )
            for r in rows
        ])
        portfolio.save()  # persist auto-name and bump updated_at for cached payloads

    note = f"Imported {len(rows)} holdings into {portfolio.name}."
    if auto_named:
        note += " Portfolio name was read from the report."
    if skipped:
        note += f" ({skipped} non-position lines skipped.)"
    messages.success(request, note)
    return redirect(_portfolio_url(portfolio))


@login_required(login_url="login")
@require_POST
def portfolio_wacc_import(request):
    """Import cost basis from the broker 'My WACC' report (CSV / Excel / PDF).

    Replaces the portfolio's ``HoldingCost`` rows and matches them to the current
    holdings by symbol. Independent of the 'My Shares' import, so cost basis
    survives a holdings re-upload.
    """
    from core_analysis.models import HoldingCost

    portfolio = _get_mutation_portfolio(request)
    if portfolio is None:
        messages.error(request, "Portfolio not found.")
        return redirect(_portfolio_url(_get_selected_portfolio(request)))

    upload = request.FILES.get("file")
    if not upload:
        messages.error(request, "Please choose your 'My WACC' report to upload.")
        return redirect(_portfolio_url(portfolio))
    upload_error = _upload_validation_error(upload, "WACC")
    if upload_error:
        messages.error(request, upload_error)
        return redirect(_portfolio_url(portfolio))

    try:
        rows, skipped = parse_wacc_report(upload)
    except Exception:
        logger.exception("WACC parse failed for user %s", request.user.id)
        messages.error(
            request,
            "Could not read that file — upload the broker 'My WACC' report (CSV, Excel or PDF).",
        )
        return redirect(_portfolio_url(portfolio))

    if not rows:
        messages.error(request, "No WACC rows found in that file. Check the format and try again.")
        return redirect(_portfolio_url(portfolio))

    with transaction.atomic():
        portfolio = _locked_active_portfolio(request.user, portfolio.id)
        if portfolio is None:
            messages.error(request, "Portfolio not found or no longer active.")
            return redirect(_portfolio_url(_get_selected_portfolio(request)))
        held = set(portfolio.holdings.values_list("symbol", flat=True))
        portfolio.costs.all().delete()
        HoldingCost.objects.bulk_create([
            HoldingCost(
                portfolio=portfolio,
                symbol=r["symbol"],
                wacc_rate=r["wacc_rate"],
                quantity=r["quantity"] or 0,
                total_cost=r["total_cost"],
                modified=r["modified"] or "",
            )
            for r in rows
        ])
        portfolio.save()  # bump updated_at so cached payloads invalidate

    matched = sum(1 for r in rows if r["symbol"] in held)
    note = f"Imported cost basis for {len(rows)} scrips into {portfolio.name} — {matched} matched your holdings."
    missing = len(held) - matched
    if missing > 0:
        note += f" ({missing} holding(s) still without cost.)"
    if skipped:
        note += f" ({skipped} non-data lines skipped.)"
    messages.success(request, note)
    return redirect(_portfolio_url(portfolio))


@login_required(login_url="login")
@require_POST
def portfolio_ledger_import(request):
    """Merge a broker account statement into the selected private portfolio."""
    from core_analysis.services.portfolio_ledger import import_parsed_ledger

    portfolio = _get_mutation_portfolio(request)
    if portfolio is None:
        messages.error(request, "Portfolio not found.")
        return redirect(_portfolio_url(_get_selected_portfolio(request)))

    upload = request.FILES.get("file")
    if not upload:
        messages.error(request, "Please choose a broker Account Statement to upload.")
        return redirect(_portfolio_url(portfolio))
    upload_error = _upload_validation_error(upload, "broker ledger")
    if upload_error:
        messages.error(request, upload_error)
        return redirect(_portfolio_url(portfolio))

    try:
        parsed, file_sha256 = parse_broker_ledger_report(upload)
    except Exception:
        logger.exception("broker ledger parse failed for user %s", request.user.id)
        messages.error(
            request,
            "Could not read that file — upload the broker Account Statement (CSV, Excel or PDF).",
        )
        return redirect(_portfolio_url(portfolio))

    try:
        with transaction.atomic():
            portfolio = _locked_active_portfolio(request.user, portfolio.id)
            if portfolio is None:
                messages.error(request, "Portfolio not found or no longer active.")
                return redirect(_portfolio_url(_get_selected_portfolio(request)))
            result = import_parsed_ledger(
                portfolio,
                parsed,
                source_name=upload.name or "Broker Account Statement",
                file_sha256=file_sha256,
            )
    except IntegrityError:
        logger.exception("broker ledger import conflicted for user %s", request.user.id)
        messages.error(request, "That ledger overlaps an import being processed. Please retry.")
        return redirect(_portfolio_url(portfolio))

    if result["already_imported"]:
        messages.info(request, "This exact broker ledger file was already imported; no rows changed.")
    else:
        note = f"Imported {result['imported']} broker ledger transaction(s) into {portfolio.name}."
        if result["duplicates"]:
            note += f" Skipped {result['duplicates']} duplicate row(s)."
        if parsed["warnings"]:
            note += f" Flagged {len(parsed['warnings'])} reconciliation warning(s)."
        messages.success(request, note)
    return redirect(_portfolio_url(portfolio))


@login_required(login_url="login")
@require_POST
def portfolio_clear(request):
    """Wipe the user's holdings (keeps the empty portfolio)."""
    portfolio = _get_mutation_portfolio(request)
    if portfolio is None:
        messages.error(request, "Portfolio not found.")
        return redirect(_portfolio_url(_get_selected_portfolio(request)))
    with transaction.atomic():
        portfolio = _locked_active_portfolio(request.user, portfolio.id)
        if portfolio is None:
            messages.error(request, "Portfolio not found or no longer active.")
            return redirect(_portfolio_url(_get_selected_portfolio(request)))
        portfolio.holdings.all().delete()
        portfolio.save()
    messages.success(request, f"Holdings cleared for {portfolio.name}.")
    return redirect(_portfolio_url(portfolio))
